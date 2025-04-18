#!/usr/bin/env python3
"""
Script to generate sample spreadsheet data for testing.
"""

import os
import argparse
import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def generate_sample_data(rows=50, cols=20, output_file="Data/samples/sample_data.xlsx"):
    """Generate sample spreadsheet with typical business data."""
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"
    
    # Define data types to use
    data_types = ["header", "int", "float", "date", "text", "formula", "percentage", "currency"]
    
    # Create header row with bold formatting
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    header_border = Border(bottom=Side(style='medium'))
    
    # Define column headers
    headers = [
        "ID", "Date", "Customer", "Region", "Product", "Quantity", 
        "Unit Price", "Total", "Tax Rate", "Tax Amount", "Grand Total",
        "Payment Status", "Payment Date", "Invoice #", "Notes"
    ]
    
    # Extend headers if needed
    while len(headers) < cols:
        headers.append(f"Extra Col {len(headers) + 1}")
    
    # Set headers
    for col in range(1, min(cols, len(headers)) + 1):
        cell = ws.cell(row=1, column=col, value=headers[col-1])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')
    
    # Generate data rows
    regions = ["North", "South", "East", "West", "Central"]
    products = ["Widget A", "Widget B", "Service C", "Premium D", "Basic E"]
    statuses = ["Paid", "Pending", "Overdue", "Cancelled"]
    
    for row in range(2, rows + 2):
        # Create some empty rows occasionally
        if random.random() < 0.05 and row > 5:
            continue
            
        # Basic data
        ws.cell(row=row, column=1, value=row-1)  # ID
        date_val = datetime.datetime.now() - datetime.timedelta(days=random.randint(0, 365))
        ws.cell(row=row, column=2, value=date_val)  # Date
        ws.cell(row=row, column=3, value=f"Customer-{random.randint(1000, 9999)}")  # Customer
        ws.cell(row=row, column=4, value=random.choice(regions))  # Region
        ws.cell(row=row, column=5, value=random.choice(products))  # Product
        
        # Numeric data
        quantity = random.randint(1, 100)
        ws.cell(row=row, column=6, value=quantity)  # Quantity
        
        unit_price = round(random.uniform(10, 500), 2)
        price_cell = ws.cell(row=row, column=7, value=unit_price)  # Unit Price
        price_cell.number_format = '"$"#,##0.00'
        
        # Formulas
        total_cell = ws.cell(row=row, column=8)  # Total
        total_cell.value = f"=F{row}*G{row}"
        total_cell.number_format = '"$"#,##0.00'
        
        tax_rate = round(random.uniform(0.05, 0.25), 2)
        tax_rate_cell = ws.cell(row=row, column=9, value=tax_rate)  # Tax Rate
        tax_rate_cell.number_format = '0.00%'
        
        tax_cell = ws.cell(row=row, column=10)  # Tax Amount
        tax_cell.value = f"=H{row}*I{row}"
        tax_cell.number_format = '"$"#,##0.00'
        
        grand_total_cell = ws.cell(row=row, column=11)  # Grand Total
        grand_total_cell.value = f"=H{row}+J{row}"
        grand_total_cell.number_format = '"$"#,##0.00'
        
        # Other data
        status = random.choice(statuses)
        status_cell = ws.cell(row=row, column=12, value=status)  # Payment Status
        if status == "Paid":
            status_cell.font = Font(color="006100")
        elif status == "Overdue":
            status_cell.font = Font(color="FF0000")
        
        # Payment date (only if paid)
        if status == "Paid":
            payment_date = date_val + datetime.timedelta(days=random.randint(1, 30))
            ws.cell(row=row, column=13, value=payment_date)  # Payment Date
        
        # Invoice number
        ws.cell(row=row, column=14, value=f"INV-{date_val.year}{date_val.month:02d}-{row-1:04d}")  # Invoice #
        
        # Notes (occasionally)
        if random.random() < 0.3:
            notes = random.choice([
                "Priority customer",
                "Discount applied",
                "Repeat order",
                "New customer",
                "Special handling required",
                "Seasonal promotion"
            ])
            ws.cell(row=row, column=15, value=notes)  # Notes
    
    # Create a summary section at the bottom with merged cells
    summary_row = rows + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=3)
    
    # Total calculations
    ws.cell(row=summary_row + 1, column=1, value="Total Quantity:")
    ws.cell(row=summary_row + 1, column=2, value=f"=SUM(F2:F{rows+1})").font = Font(bold=True)
    
    ws.cell(row=summary_row + 2, column=1, value="Total Revenue:")
    revenue_cell = ws.cell(row=summary_row + 2, column=2, value=f"=SUM(H2:H{rows+1})")
    revenue_cell.font = Font(bold=True)
    revenue_cell.number_format = '"$"#,##0.00'
    
    ws.cell(row=summary_row + 3, column=1, value="Average Order:")
    avg_cell = ws.cell(row=summary_row + 3, column=2, value=f"=AVERAGE(H2:H{rows+1})")
    avg_cell.font = Font(bold=True)
    avg_cell.number_format = '"$"#,##0.00'
    
    # Add borders to summary section
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in range(summary_row, summary_row + 4):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border
    
    # Adjust column widths
    for col in range(1, cols + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Sample data generated and saved to {output_file}")

def main():
    """Parse arguments and generate sample data."""
    parser = argparse.ArgumentParser(description="Generate sample spreadsheet data for testing")
    parser.add_argument("--rows", type=int, default=50, help="Number of data rows")
    parser.add_argument("--cols", type=int, default=15, help="Number of columns")
    parser.add_argument("--output", default="Data/samples/sample_data.xlsx", help="Output file path")
    
    args = parser.parse_args()
    generate_sample_data(args.rows, args.cols, args.output)

if __name__ == "__main__":
    main() 