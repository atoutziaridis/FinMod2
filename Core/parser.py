"""Spreadsheet parsing module for extracting data and formatting."""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from typing import Dict, List, Tuple, Any, Optional, Set, Union
import numpy as np
from dataclasses import dataclass, field

from Core.utils.constants import DataType, FormatType
from Core.utils.helpers import infer_data_type

@dataclass
class CellFormat:
    """Class to store cell formatting information."""
    formats: Set[FormatType] = field(default_factory=set)
    bg_color: Optional[str] = None
    font_color: Optional[str] = None
    borders: Set[str] = field(default_factory=set)
    
    def __str__(self) -> str:
        """String representation of cell format."""
        parts = []
        for fmt in self.formats:
            parts.append(fmt.name.lower())
        if self.bg_color:
            parts.append(f"bg:{self.bg_color}")
        if self.font_color:
            parts.append(f"fg:{self.font_color}")
        if self.borders:
            parts.append(f"borders:{','.join(self.borders)}")
        return " ".join(parts)

@dataclass
class Cell:
    """Class to store cell information."""
    value: Any
    data_type: DataType
    format: CellFormat
    address: str
    
    def __str__(self) -> str:
        """String representation of cell."""
        return f"{self.address}:{self.value}({self.data_type.name})"

@dataclass
class SheetMatrix:
    """Class to store sheet data in matrix form."""
    matrix: List[List[Optional[Cell]]]
    merged_cells: List[Tuple[int, int, int, int]]  # (min_row, min_col, max_row, max_col)
    sheet_name: str
    max_row: int
    max_col: int
    
    def get_cell(self, row: int, col: int) -> Optional[Cell]:
        """Get cell at specified row and column."""
        if 1 <= row <= self.max_row and 1 <= col <= self.max_col:
            return self.matrix[row-1][col-1]
        return None

class SpreadsheetParser:
    """Parser for extracting data and formatting from Excel spreadsheets."""
    
    def __init__(self, filepath: str):
        """Initialize with filepath to Excel spreadsheet."""
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
        
    def parse_sheet(self, sheet_name: Optional[str] = None) -> SheetMatrix:
        """Parse specified sheet or active sheet if not specified."""
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
            sheet_name = sheet.title
            
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Initialize matrix
        matrix = [[None for _ in range(max_col)] for _ in range(max_row)]
        
        # Process cells
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell_address = f"{get_column_letter(col)}{row}"
                
                # Extract value and infer data type
                value = cell.value
                data_type = infer_data_type(value)
                
                # Extract formatting
                cell_format = self._extract_format(cell)
                
                # Create Cell object
                matrix[row-1][col-1] = Cell(
                    value=value,
                    data_type=data_type,
                    format=cell_format,
                    address=cell_address
                )
        
        # Extract merged cells
        merged_cells = []
        for merged_range in sheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            merged_cells.append((min_row, min_col, max_row, max_col))
        
        return SheetMatrix(
            matrix=matrix,
            merged_cells=merged_cells,
            sheet_name=sheet_name,
            max_row=max_row,
            max_col=max_col
        )
    
    def _extract_format(self, cell: openpyxl.cell.Cell) -> CellFormat:
        """Extract formatting information from cell."""
        fmt = CellFormat()
        
        # Font
        if cell.font:
            if cell.font.bold:
                fmt.formats.add(FormatType.BOLD)
            if cell.font.italic:
                fmt.formats.add(FormatType.ITALIC)
            if cell.font.underline:
                fmt.formats.add(FormatType.UNDERLINE)
            if cell.font.strike:
                fmt.formats.add(FormatType.STRIKETHROUGH)
            if cell.font.color:
                fmt.font_color = cell.font.color.rgb
        
        # Fill
        if cell.fill and cell.fill.fill_type == 'solid' and cell.fill.start_color:
            fmt.bg_color = cell.fill.start_color.rgb
            fmt.formats.add(FormatType.HIGHLIGHTED)
        
        # Borders
        if cell.border:
            borders = []
            if cell.border.left and cell.border.left.style:
                borders.append('left')
            if cell.border.right and cell.border.right.style:
                borders.append('right')
            if cell.border.top and cell.border.top.style:
                borders.append('top')
            if cell.border.bottom and cell.border.bottom.style:
                borders.append('bottom')
            if borders:
                fmt.borders = set(borders)
                fmt.formats.add(FormatType.BORDERED)
        
        # Alignment
        if cell.alignment:
            if cell.alignment.horizontal == 'center':
                fmt.formats.add(FormatType.CENTER_ALIGNED)
            elif cell.alignment.horizontal == 'left':
                fmt.formats.add(FormatType.LEFT_ALIGNED)
            elif cell.alignment.horizontal == 'right':
                fmt.formats.add(FormatType.RIGHT_ALIGNED)
        
        return fmt
    
    def get_format_dict(self, sheet_matrix: SheetMatrix) -> Dict[Tuple[int, int], DataType]:
        """Create a format dictionary mapping (row, col) to value type."""
        format_dict = {}
        for row in range(1, sheet_matrix.max_row + 1):
            for col in range(1, sheet_matrix.max_col + 1):
                cell = sheet_matrix.get_cell(row, col)
                if cell and cell.value is not None:
                    format_dict[(row, col)] = cell.data_type
        return format_dict 