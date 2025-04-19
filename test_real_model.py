#!/usr/bin/env python3
"""
Test script to analyze and debug the real Excel model implementation.
"""

import os
import json
import argparse
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict

from Core.parser import SpreadsheetParser
from Core.encoder import SpreadsheetEncoder
from Core.compressor import SheetCompressor
from Core.utils.helpers import estimate_tokens, detect_number_format_string

def analyze_sheet(sheet_matrix, sheet_name):
    """Analyze a single sheet and print statistics."""
    print(f"\nAnalyzing sheet: {sheet_name}")
    print(f"Dimensions: {sheet_matrix.max_row}x{sheet_matrix.max_col}")
    
    # Count non-empty cells
    non_empty_count = sum(1 for row in range(1, sheet_matrix.max_row + 1) 
                         for col in range(1, sheet_matrix.max_col + 1) 
                         if sheet_matrix.get_cell(row, col) is not None and 
                            sheet_matrix.get_cell(row, col).value is not None)
    
    print(f"Non-empty cells: {non_empty_count}")
    
    # Count different data types
    data_types = {}
    for row in range(1, sheet_matrix.max_row + 1):
        for col in range(1, sheet_matrix.max_col + 1):
            cell = sheet_matrix.get_cell(row, col)
            if cell and cell.value is not None:
                data_type = cell.data_type.name
                data_types[data_type] = data_types.get(data_type, 0) + 1
    
    print("\nData type distribution:")
    for dtype, count in data_types.items():
        print(f"{dtype}: {count}")
    
    return {
        "dimensions": f"{sheet_matrix.max_row}x{sheet_matrix.max_col}",
        "non_empty_cells": non_empty_count,
        "data_types": data_types
    }

def test_compression_modules(sheet_matrix, sheet_name, output_dir, token_limit=8000, skip_if_larger_than=None):
    """Test different compression module combinations."""
    results = {}
    encoder = SpreadsheetEncoder(max_tokens=token_limit)
    
    # Check if sheet is too large to process
    estimated_size = sheet_matrix.max_row * sheet_matrix.max_col * 2  # Rough estimate 
    if skip_if_larger_than and estimated_size > skip_if_larger_than:
        print(f"Sheet {sheet_name} is too large (est. {estimated_size} tokens), skipping detailed analysis")
        
        # Just do a quick analysis with the most promising approach (Module 3)
        try:
            print("Trying Module 3 (Data-Format-Aware Aggregation) for large sheet...")
            encoder.use_inverted_index = False
            encoder.use_format_aggregation = True
            
            try:
                m3_encoded = encoder.encode_sheet(sheet_matrix)
                m3_tokens = estimate_tokens(m3_encoded)
                
                # Save only the final output
                final_file = os.path.join(output_dir, f"{sheet_name}_combined_all_modules.json")
                with open(final_file, 'w') as f:
                    f.write(m3_encoded)
                print(f"Successfully encoded large sheet with Module 3: {m3_tokens} tokens")
                
                return {
                    "module3": {
                        "tokens": m3_tokens,
                        "file": final_file,
                        "compression_ratio": "N/A (large sheet)"
                    }
                }
            except ValueError:
                print("Module 3 failed, sheet is too large for standard encoding")
                return {"error": "Sheet too large for encoding"}
                
        except Exception as e:
            print(f"Error processing large sheet: {str(e)}")
            return {"error": str(e)}
    
    # Original (no compression) - just for comparison, don't save
    print("\nTesting vanilla encoding (no compression)...")
    try:
        encoder.use_inverted_index = False
        encoder.use_format_aggregation = False
        
        vanilla_encoded = encoder.encode_sheet(sheet_matrix)
        vanilla_tokens = estimate_tokens(vanilla_encoded)
        print(f"Vanilla encoding token count: {vanilla_tokens}")
        
        results["vanilla"] = {
            "tokens": vanilla_tokens,
            "compression_ratio": 1.0  # Baseline
        }
    except ValueError as e:
        print(f"Error with vanilla encoding: {str(e)}")
        results["vanilla"] = {"error": str(e)}
    
    # Test Module 1: Structural Anchor Compression
    print("\nTesting Module 1: Structural Anchor Compression...")
    try:
        compressor = SheetCompressor()
        compressed_m1 = compressor.compress(sheet_matrix)
        
        # Encode compressed sheet
        encoder.use_inverted_index = False
        encoder.use_format_aggregation = False
        m1_encoded = encoder.encode_sheet(compressed_m1)
        m1_tokens = estimate_tokens(m1_encoded)
        
        compression_ratio = m1_tokens / vanilla_tokens if vanilla_tokens > 0 else 0
        print(f"Module 1 token count: {m1_tokens}")
        print(f"Module 1 compression ratio: {1/compression_ratio:.2f}x")
        
        results["module1"] = {
            "tokens": m1_tokens,
            "compression_ratio": 1/compression_ratio
        }
    except Exception as e:
        print(f"Error with Module 1: {str(e)}")
        results["module1"] = {"error": str(e)}
    
    # Test Module 2: Inverted Index Translation
    print("\nTesting Module 2: Inverted Index Translation...")
    try:
        encoder.use_inverted_index = True
        encoder.use_format_aggregation = False
        m2_encoded = encoder.encode_sheet(sheet_matrix)
        m2_tokens = estimate_tokens(m2_encoded)
        
        compression_ratio = m2_tokens / vanilla_tokens if vanilla_tokens > 0 else 0
        print(f"Module 2 token count: {m2_tokens}")
        print(f"Module 2 compression ratio: {1/compression_ratio:.2f}x")
        
        results["module2"] = {
            "tokens": m2_tokens,
            "compression_ratio": 1/compression_ratio
        }
    except Exception as e:
        print(f"Error with Module 2: {str(e)}")
        results["module2"] = {"error": str(e)}
    
    # Test Module 3: Data-Format-Aware Aggregation
    print("\nTesting Module 3: Data-Format-Aware Aggregation...")
    try:
        encoder.use_inverted_index = False
        encoder.use_format_aggregation = True
        m3_encoded = encoder.encode_sheet(sheet_matrix)
        m3_tokens = estimate_tokens(m3_encoded)
        
        compression_ratio = m3_tokens / vanilla_tokens if vanilla_tokens > 0 else 0
        print(f"Module 3 token count: {m3_tokens}")
        print(f"Module 3 compression ratio: {1/compression_ratio:.2f}x")
        
        results["module3"] = {
            "tokens": m3_tokens,
            "compression_ratio": 1/compression_ratio
        }
    except Exception as e:
        print(f"Error with Module 3: {str(e)}")
        results["module3"] = {"error": str(e)}
    
    # Test Combined: All three modules
    print("\nTesting Combined: All three modules...")
    try:
        # First apply structural compression
        compressor = SheetCompressor()
        compressed_combined = compressor.compress(sheet_matrix)
        
        # Then apply inverted index and format aggregation
        encoder.use_inverted_index = True
        encoder.use_format_aggregation = True
        combined_encoded = encoder.encode_sheet(compressed_combined)
        combined_tokens = estimate_tokens(combined_encoded)
        
        # Save only the final combined output
        final_file = os.path.join(output_dir, f"{sheet_name}_combined_all_modules.json")
        with open(final_file, 'w') as f:
            f.write(combined_encoded)
        
        compression_ratio = combined_tokens / vanilla_tokens if vanilla_tokens > 0 else 0
        print(f"Combined token count: {combined_tokens}")
        print(f"Combined compression ratio: {1/compression_ratio:.2f}x")
        
        results["combined"] = {
            "tokens": combined_tokens,
            "file": final_file,
            "compression_ratio": 1/compression_ratio
        }
    except Exception as e:
        print(f"Error with Combined approach: {str(e)}")
        results["combined"] = {"error": str(e)}
    
    return results

def compress_with_best_method(sheet_matrix, output_dir, filename_prefix, max_tokens):
    """Compress the sheet using the best available method."""
    results = {}
    
    # Test all compression modules
    compression_results = test_compression_modules(
        sheet_matrix, 
        filename_prefix, 
        output_dir,
        token_limit=max_tokens
    )
    
    # Determine best approach
    best_approach = None
    best_ratio = float('-inf')
    
    for module, result in compression_results.items():
        if "error" in result:
            continue
        if "compression_ratio" in result and result["compression_ratio"] > best_ratio:
            best_ratio = result["compression_ratio"]
            best_approach = module
    
    if best_approach:
        results["method"] = best_approach
        results["compression_ratio"] = best_ratio
        results["output_file"] = compression_results[best_approach]["file"]
        results["tokens"] = compression_results[best_approach]["tokens"]
    else:
        results["error"] = "No valid compression method found"
    
    return results

def process_sheet(sheet_name: str, model_path: str, output_dir: str, token_limit: int) -> Dict:
    """Process a single sheet and return results."""
    try:
        print(f"Processing sheet: {sheet_name}")
        parser = SpreadsheetParser(model_path)
        sheet_matrix = parser.parse_sheet(sheet_name)
        
        # Get basic statistics
        sheet_stats = analyze_sheet(sheet_matrix, sheet_name)
        
        # Use the best compression method automatically
        compression_results = compress_with_best_method(
            sheet_matrix,
            output_dir,
            filename_prefix=sheet_name,
            max_tokens=token_limit
        )
        
        # Update sheet statistics with compression results
        sheet_stats.update({
            "compression_method": compression_results["method"],
            "compression_ratio": compression_results["compression_ratio"],
            "output_file": compression_results["output_file"],
            "tokens": compression_results["tokens"]
        })
        
        print(f"Completed sheet {sheet_name}")
        return {sheet_name: sheet_stats}
        
    except Exception as e:
        print(f"Error processing sheet {sheet_name}: {str(e)}")
        return {sheet_name: {"error": str(e)}}

def main():
    """Run analysis on the real Excel model."""
    # Parse command line arguments
    parser = argparse.ArgumentParser(description="Test SpreadsheetLLM Framework on real Excel models")
    parser.add_argument("--file", default="Data/Real Data/Miscounted Cash Flow.ExcelModel.xlsx", 
                      help="Path to Excel file")
    parser.add_argument("--output-dir", default="output", help="Output directory")
    parser.add_argument("--sheet", help="Specific sheet to process (omit to process all sheets)")
    parser.add_argument("--token-limit", type=int, default=8000, help="Token limit for encoding")
    parser.add_argument("--parallel", type=int, default=4, help="Number of parallel processes")
    args = parser.parse_args()
    
    # Set up variables
    model_path = args.file
    output_dir = args.output_dir
    token_limit = args.token_limit
    specific_sheet = args.sheet
    parallel = args.parallel
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    if not os.path.exists(model_path):
        print(f"Error: File not found at {model_path}")
        return
    
    print(f"Analyzing real model: {model_path}")
    
    # Load workbook to get sheet names
    wb = load_workbook(model_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    if specific_sheet:
        if specific_sheet in sheet_names:
            sheet_names = [specific_sheet]
            print(f"\nProcessing only sheet: {specific_sheet}")
        else:
            print(f"Error: Sheet '{specific_sheet}' not found in the workbook")
            print(f"Available sheets: {', '.join(sheet_names)}")
            return
    else:
        print(f"\nFound {len(sheet_names)} sheets:")
        for name in sheet_names:
            print(f"- {name}")
    
    # Process sheets in parallel
    from multiprocessing import Pool
    
    # Create a pool of workers
    with Pool(processes=parallel) as pool:
        # Prepare arguments for each sheet
        args_list = [(name, model_path, output_dir, token_limit) for name in sheet_names]
        
        # Process sheets in parallel
        results = pool.starmap(process_sheet, args_list)
    
    # Combine results
    all_sheets_results = {}
    for result in results:
        all_sheets_results.update(result)
    
    # Save overall results
    stats_filename = os.path.join(output_dir, "compression_results.json")
    with open(stats_filename, 'w') as f:
        json.dump(all_sheets_results, f, indent=2)
    print(f"\nCompression results saved to: {stats_filename}")
    
    print("\nAnalysis complete!")

if __name__ == "__main__":
    main() 